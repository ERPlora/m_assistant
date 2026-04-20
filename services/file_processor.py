"""
File processing for AI Assistant.

Combined strategy:
1. PDF with text -> extract text via pdfplumber (cheap, fast)
2. PDF scanned/image-only -> convert to PNG via PyMuPDF, send as base64 to Vision
3. Images (jpg, png, webp) -> read as base64, send to Vision
4. Text files (txt, csv) -> read as text directly
5. Office docs (xlsx, docx) -> extract text where possible

Returns a list of multimodal content parts for the LLM:
[{"type": "input_text", "text": "..."}, {"type": "input_image", "image_url": "data:image/png;base64,..."}]
"""

from __future__ import annotations

import base64
import io
import logging
from typing import Any

logger = logging.getLogger(__name__)

# Max file size: 10 MB
MAX_FILE_SIZE = 10 * 1024 * 1024

# Supported MIME types
SUPPORTED_TYPES = {
    # Images -> send as Vision
    "image/jpeg", "image/png", "image/webp", "image/gif",
    # PDFs -> text extract or convert to image
    "application/pdf",
    # Text -> read directly
    "text/plain", "text/csv",
    # Office -> best effort text extraction
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # xlsx
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",  # docx
}


async def process_file(
    file_bytes: bytes,
    filename: str,
    content_type: str) -> list[dict[str, Any]]:
    """
    Process an uploaded file and return multimodal content parts.

    Returns list of:
    - {"type": "input_text", "text": "extracted text..."}
    - {"type": "input_image", "image_url": "data:image/png;base64,..."}
    """
    if len(file_bytes) > MAX_FILE_SIZE:
        return [{"type": "input_text", "text": f"[File '{filename}' too large: {len(file_bytes) / 1024 / 1024:.1f} MB, max 10 MB]"}]

    if content_type not in SUPPORTED_TYPES:
        return [{"type": "input_text", "text": f"[Unsupported file type: {content_type}]"}]

    # Route by type
    if content_type.startswith("image/"):
        return _process_image(file_bytes, filename, content_type)
    elif content_type == "application/pdf":
        return _process_pdf(file_bytes, filename)
    elif content_type in ("text/plain", "text/csv"):
        return _process_text(file_bytes, filename)
    elif "spreadsheetml" in content_type:
        return _process_xlsx(file_bytes, filename)
    elif "wordprocessingml" in content_type:
        return _process_docx(file_bytes, filename)

    return [{"type": "input_text", "text": f"[Could not process '{filename}']"}]


def _process_image(file_bytes: bytes, filename: str, content_type: str) -> list[dict]:
    """Images -> base64 for Vision."""
    b64 = base64.b64encode(file_bytes).decode()
    mime = content_type or "image/png"
    return [
        {"type": "input_text", "text": f"[Attached image: {filename}]"},
        {"type": "input_image", "image_url": f"data:{mime};base64,{b64}"},
    ]


def _process_pdf(file_bytes: bytes, filename: str) -> list[dict]:
    """
    PDF: try text extraction first (pdfplumber), fall back to image (pymupdf).
    """
    # Strategy 1: Extract text with pdfplumber
    text = _extract_pdf_text(file_bytes)
    if text and len(text.strip()) > 50:
        # Enough text found -- send as text (cheap)
        return [{"type": "input_text", "text": f"[Document: {filename}]\n\n{text}"}]

    # Strategy 2: Convert to images with PyMuPDF (scanned PDF)
    images = _pdf_to_images(file_bytes)
    if images:
        parts: list[dict] = [{"type": "input_text", "text": f"[Scanned document: {filename}, {len(images)} page(s)]"}]
        for img_bytes in images[:5]:  # Max 5 pages
            b64 = base64.b64encode(img_bytes).decode()
            parts.append({"type": "input_image", "image_url": f"data:image/png;base64,{b64}"})
        return parts

    return [{"type": "input_text", "text": f"[Could not read PDF: {filename}]"}]


def _extract_pdf_text(file_bytes: bytes) -> str:
    """Extract text from PDF using pdfplumber."""
    try:
        import pdfplumber
        pages_text = []
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for i, page in enumerate(pdf.pages[:20]):  # Max 20 pages
                text = page.extract_text() or ""
                # Also try to extract tables
                tables = page.extract_tables()
                table_text = ""
                if tables:
                    for table in tables:
                        for row in table:
                            cells = [str(c or "").strip() for c in row]
                            table_text += " | ".join(cells) + "\n"

                page_content = text
                if table_text:
                    page_content += "\n\n[Table]\n" + table_text

                if page_content.strip():
                    pages_text.append(f"--- Page {i + 1} ---\n{page_content}")

        return "\n\n".join(pages_text)
    except ImportError:
        logger.warning("pdfplumber not installed -- cannot extract PDF text")
        return ""
    except Exception as e:
        logger.warning("PDF text extraction failed: %s", e)
        return ""


def _pdf_to_images(file_bytes: bytes) -> list[bytes]:
    """Convert PDF pages to PNG images using PyMuPDF."""
    try:
        import fitz  # PyMuPDF
        images = []
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        for i, page in enumerate(doc):
            if i >= 5:  # Max 5 pages for Vision
                break
            # Render at 150 DPI (good balance quality/size)
            mat = fitz.Matrix(150 / 72, 150 / 72)
            pix = page.get_pixmap(matrix=mat)
            images.append(pix.tobytes("png"))
        doc.close()
        return images
    except ImportError:
        logger.warning("PyMuPDF not installed -- cannot convert PDF to images")
        return []
    except Exception as e:
        logger.warning("PDF to image conversion failed: %s", e)
        return []


def _process_text(file_bytes: bytes, filename: str) -> list[dict]:
    """Plain text / CSV -> read directly."""
    try:
        text = file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1", errors="replace")

    # Truncate very long files
    if len(text) > 50_000:
        text = text[:50_000] + f"\n\n[... truncated, {len(file_bytes)} bytes total]"

    return [{"type": "input_text", "text": f"[File: {filename}]\n\n{text}"}]


def _process_xlsx(file_bytes: bytes, filename: str) -> list[dict]:
    """Excel -> extract text from cells."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheets_text = []
        for sheet_name in wb.sheetnames[:5]:  # Max 5 sheets
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(max_row=200, values_only=True):
                cells = [str(c or "").strip() for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                sheets_text.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
        wb.close()
        text = "\n\n".join(sheets_text)
        if text:
            return [{"type": "input_text", "text": f"[Excel: {filename}]\n\n{text}"}]
    except ImportError:
        logger.warning("openpyxl not installed -- cannot read Excel files")
    except Exception as e:
        logger.warning("Excel extraction failed: %s", e)

    return [{"type": "input_text", "text": f"[Could not read Excel: {filename}]"}]


def _process_docx(file_bytes: bytes, filename: str) -> list[dict]:
    """Word -> extract paragraphs."""
    try:
        import docx
        doc = docx.Document(io.BytesIO(file_bytes))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n".join(paragraphs)
        if text:
            return [{"type": "input_text", "text": f"[Document: {filename}]\n\n{text}"}]
    except ImportError:
        logger.warning("python-docx not installed -- cannot read Word files")
    except Exception as e:
        logger.warning("Word extraction failed: %s", e)

    return [{"type": "input_text", "text": f"[Could not read Word document: {filename}]"}]
